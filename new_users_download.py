# -*- coding: cp1255 -*-   # sets the coding to hebrew and not gibrish

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from datetime import datetime

from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

CHROME_DRIVER_PATH = "C:\\Users\\ronke\\Desktop\\פרויקט גמר\\chromedriver.exe"
DEFAULT_URL = "https://www.camoni.co.il/411788/"
DOWNLOAD_PATH = "C:\\Users\\ronke\\Desktop\\פרויקט גמר\\parsing\\users_download_new_try\\"
LOG_FILE_PATH = DOWNLOAD_PATH + "log_file_users_download_" + str(datetime.now().strftime("%d-%m-%y")) + ".txt"
users_count = 0

USERS_OLD_FILE = r'C:\Users\ronke\Desktop\פרויקט גמר\parsing\DATA\Camoni_USERS__24-Apr-2022-09_10.xlsx'
wb = openpyxl.load_workbook(USERS_OLD_FILE)
ws = wb.get_sheet_by_name('Camoni')
# for every row in wb, get the url and add to the list
URL_list = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=3, max_col=3):
    for cell in row:
        cell_list = cell.value.split("\"")
        url = cell_list[1]
        url = url.replace(" ", "")
        URL_list.append(url)


def double_print(my_str):
    global log_file
    print(my_str)
    log_file.write(my_str)


log_file = open(LOG_FILE_PATH, 'w')
double_print("\n" + str(datetime.now().strftime("%d/%m/%y %H:%M:%S")) + " - Starting getting users file")
double_print(
    "-----------------------------------------------------------------------------------------------------------------")

chrome_options = Options()
chrome_options.add_argument("--headless")
# browser = webdriver.Chrome(executable_path=CHROME_DRIVER_PATH, options=chrome_options)
browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

for url in URL_list:
    try:
        browser.get(url)
        content = browser.page_source
        file_name = url.split('/')[-1]
        with open(DOWNLOAD_PATH + file_name + ".html", "w", encoding='utf-8') as file:
            file.write(str(content))
            double_print(
                str(datetime.now().strftime("%d/%m/%y - %H:%M:%S")) + " - created successfully file: " + file_name)

        users_count += 1
        # time.sleep(0.5)

    except Exception as e:
        double_print(file_name + " failed because: " + str(e))

browser.close()

double_print(
    "-----------------------------------------------------------------------------------------------------------------")
double_print(
    "\n" + str(datetime.now().strftime("%d/%m/%y - %H:%M:%S")) + " - Downloaded " + str(users_count) + " files!")
log_file.close()
